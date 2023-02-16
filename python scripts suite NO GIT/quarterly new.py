
# u prvoj koloni mora da bude customer

# sheet-ovi se moraju zvati:
#   DCS Deployments
#   DCS Inventory

# libraries needed for exporting to excel
# pip install pandas
# pip install openpyxl

import os
import shutil
from time import sleep
import openpyxl.worksheet.table
import pandas
from openpyxl.utils import get_column_letter

print("importing customer configuration")

# import customer configuration
from customers import masterdict
# remove generic
del masterdict["generic"]

print("preparing directory structure")

# pripare directory structure
source_dir = "source"
reports_dir = "reports"
reports_dir_dcs_deployments = reports_dir + os.sep + "reports_dir_dcs_deployments"
reports_dir_dcs_inventory = reports_dir + os.sep + "reports_dir_dcs_inventory"

try:
    shutil.rmtree(reports_dir)
except:
    print("error deleting reports directory structure")
    sleep(3)

# wait for windows to finish
sleep(3)

os.mkdir(reports_dir)
os.mkdir(reports_dir_dcs_deployments)
os.mkdir(reports_dir_dcs_inventory)

print("parsing source data")

for file in os.scandir(source_dir):
    print("parsing through file: " + file.name)
    dataframes_dictionary = pandas.read_excel(file, sheet_name=None, engine='openpyxl')
    print("finished reading of file: " + file.name)

    for sheet_name, sheet_object in dataframes_dictionary.items():
        print("accessing single sheet", sheet_name)

        # set output file & directory
        if sheet_name == "DCS Deployments":
            output_directory = reports_dir_dcs_deployments
            output_filename_extension = " - Software Distribution and Installation Summary.xlsx"
        elif sheet_name == "DCS Inventory":
            output_directory = reports_dir_dcs_inventory
            output_filename_extension = " - Software Inventory Report.xlsx"
        else:
            print("wrong sheet")
            continue

        print("output directory ", output_directory)
        print("output filename extension ", output_filename_extension)

        count_row = len(sheet_object)
        count_col = len(sheet_object.columns)
        print("master sheet dimensions: ", count_row, " x ", count_col)

        # get first column name from sheet object in pandas dataframe
        first_column_name = list(sheet_object.columns)[0]
        print("first column name ", first_column_name)

        for customer in masterdict:
            # comprehension
            resultant_sheet = sheet_object[sheet_object[first_column_name] == customer]

            count_row = len(resultant_sheet)
            count_col = len(resultant_sheet.columns)
            print("          customer sheet dimensions: ", customer, " ", count_row, " x ", count_col)

            output_filename = output_directory + os.sep + customer + output_filename_extension
            resultant_sheet.to_excel(output_filename)

print("formatting excel reports")
for subdir, dirs, files in os.walk(reports_dir):
    for filename in files:
        print("formatting file: ", filename)
        filepath = subdir + os.sep + filename
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

        # remove first column
        worksheet.delete_cols(1)

        # insert table
        table_name = worksheet.title.replace(" ", "")
        table_length = worksheet.max_row
        table_width = get_column_letter(worksheet.max_column)
        tab = openpyxl.worksheet.table.Table(displayName=table_name, ref=f'A1:{table_width}{table_length}')
        style = openpyxl.worksheet.table.TableStyleInfo(name="None", showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        # autofit column width
        for column in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column)
            length = length * 1.23
            worksheet.column_dimensions[column[0].column_letter].width = length

        workbook.save(filepath)
print()
print("all reports finished")
sleep(3)
