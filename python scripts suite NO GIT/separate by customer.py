
# for one single sheet file xlsx or csv
# must have one of columns: customer of Customer or customers or Customers

# libraries needed
# pip install pandas
# pip install openpyxl

import os
import shutil
from time import sleep
import openpyxl.worksheet.table
import pandas
from openpyxl.utils import get_column_letter

print("importing customer configuration")
# import customer configuration
from customers import masterdict
# remove generic
del masterdict["generic"]

print("preparing directory structure")
source_dir = "source"
reports_dir = "reports"
try:
    shutil.rmtree(reports_dir)
except:
    print("error deleting reports directory structure")
    sleep(3)
# wait for windows to finish
sleep(3)
os.mkdir(reports_dir)

print("parsing source data")
for file in os.scandir(source_dir):

    print("parsing through file: " + file.name)
    # checking for file type
    file_extension = os.path.splitext(file)[1]
    file_name = os.path.splitext(file)[0]
    file_name = file_name.split(os.sep)
    file_name = file_name[1]

    if file_extension == '.csv':
        # returns single dataframe
        dataframe = pandas.read_csv(file)
    elif file_extension == '.xlsx':
        # returns dictionary of dataframes
        dataframe = pandas.read_excel(file, sheet_name=0, engine='openpyxl')
    else:
        raise RuntimeError('File extension not recognized')
    print("finished reading of file: " + file.name)

    # get list of column names
    list_of_column_names = list(dataframe.columns)

    # set customer of Customer or customers or Customers
    if "customer" in list_of_column_names:
        customer_column_name = "customer"
    if "Customer" in list_of_column_names:
        customer_column_name = "Customer"
    if "customers" in list_of_column_names:
        customer_column_name = "customers"
    if "Customers" in list_of_column_names:
        customer_column_name = "Customers"

    # filter out customers
    dataframe = dataframe[dataframe[customer_column_name].isin(masterdict)]

    output_filename = reports_dir + os.sep + file_name + "_output_list" + ".xlsx"
    dataframe.to_excel(output_filename)

print("formatting excel reports")
for subdir, dirs, files in os.walk(reports_dir):
    for filename in files:
        print("formatting file: ", filename)
        filepath = subdir + os.sep + filename
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

        # insert table
        table_name = worksheet.title.replace(" ", "")
        table_length = worksheet.max_row
        table_width = get_column_letter(worksheet.max_column)
        tab = openpyxl.worksheet.table.Table(displayName=table_name, ref=f'A1:{table_width}{table_length}')
        style = openpyxl.worksheet.table.TableStyleInfo(name="None", showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        workbook.save(filepath)
print()
print("all reports finished")
sleep(3)
