
# libraries needed for exporting to excel
# pip install pandas
# pip install openpyxl

import os
import csv
import re
import shutil
from collections import OrderedDict
from time import sleep
import dateutil
import datetime
import openpyxl.worksheet.table
from openpyxl.styles import PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import pandas

# hardcode project location to circumvent windows error
# os.chdir('put local link here')

# pprint !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
import pprint
pp = pprint.PrettyPrinter(indent=4)
# pp.pprint(indexed_config)
# sleep(888)

# import customer configuration
from customers import masterdict

# parse customer configuration
def parse_configuration(abbr_dict, temp_match_element, ThingName, Customer, HostName, data_dict, Installed):
    flag = 0
    for abbr_dict_key in abbr_dict:
        if flag == 1:
            break
        element_list = abbr_dict[abbr_dict_key]
        for element in element_list:
            if element == temp_match_element:
                # add to master dictionary (data_dict)
                data_dict_key = (ThingName, Customer, HostName)
                if data_dict_key not in data_dict:
                    data_dict[data_dict_key] = {}
                data_dict[data_dict_key][abbr_dict_key] = Installed
                flag = 1
                break

def create_data_frame(customers_data_dict):
    data_frame = pandas.DataFrame(customers_data_dict)
    data_frame = data_frame.transpose()
    data_frame = data_frame.sort_index(axis=1)
    data_frame.columns = data_frame.columns.droplevel()
    data_frame = data_frame.sort_index(level=[1, 2])
    data_frame.index.names = ['ThingName', 'Customer', 'HostName']
    return data_frame

print("setting files & directories")

reports_dir = "reports"
reports_dir_single_customers = reports_dir + os.sep + "single_customers_reports"
source_dir = "source"

excel_master_table = reports_dir + os.sep + "master_table_report.xlsx"
excel_multiple_sheets = reports_dir + os.sep + "single_customers_sheets.xlsx"

try:
    shutil.rmtree(reports_dir)
except:
    print("error deleting reports directory structure")
    sleep(3)
# wait for windows to finish
sleep(3)
os.mkdir(reports_dir)
os.mkdir(reports_dir_single_customers)

print("getting customers configuration")
ordered_config = masterdict

# add index for sorting configuration file
indexed_config = {}
index = 1

# only customers from parsed source data
customers_list_active = set()

# parse through customers configuration
# accessing customer
for customer_key in ordered_config:
    # setting customer dictionary
    customer_dict = ordered_config[customer_key]
    # adding indexed dictionary to customer
    indexed_config[customer_key] = {}
    # parsing through customer dictionary
    for data_type in customer_dict:
        # checking for package data dictionary
        if data_type == "package data":
            # setting package data dictionary
            packages_dict = customer_dict[data_type]
            # parsing through packages data dictionary
            for package_key in packages_dict:
                # adding index to package
                new_package_key = (index, package_key)
                package_value = packages_dict[package_key]
                # setting new key value
                indexed_config[customer_key][new_package_key] = package_value
                index += 1

##############################
# data structure for receiving data from csv file

# dictionary/list combo data structure: {
#   tuple(name1, name2, customer) : {software : date, software : date, ...}
# }
data_dict = {}
# dictionary/list combo data structure with added index: {
#   tuple(name1, name2, customer) : {tuple(index, software) : date, tuple(index, software) : date, ...}
# }
##############################

print("parsing source data")

log_line_count = 0

for file in os.scandir(source_dir):
    print("parsing through file: " + file.name)
    with open(file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            if row[0] == "ThingName":
                continue
            else:
                Customer = row[2]

                # parse config data
                if Customer in ordered_config:

                    # add customer in set of active customers
                    customers_list_active.add(Customer)

                    ThingName = row[0]
                    HostName = row[1]
                    Version = row[4]

                    Installed = row[5]
                    try:
                        # get date/time
                        Installed = dateutil.parser.parse(Installed)
                        Installed = Installed.date()
                    except:
                        pass

                    # override for dcs, core/icusn and winpass
                    # put version in field instead of date for dcs & core
                    # put date back in the field instead of empty for winpass
                    DetectedSoftware = row[3]

                    # regex_pattern_object = re.compile("DCS Agent") # will not work missing first word
                    # match_object = re.search(regex_pattern_object,DetectedSoftware)
                    # if match_object:
                    #     Installed = Version

                    regex_pattern_object = re.compile("Edge-StandardInterface-Core")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("ICUSN-Base")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

                    regex_pattern_object = re.compile("ICUSN-MUP")
                    match_object = re.search(regex_pattern_object,DetectedSoftware)
                    if match_object:
                        Installed = Version

#                    regex_pattern_object = re.compile("AE")
#                    match_object = re.search(regex_pattern_object,DetectedSoftware)
#                    if match_object:
#                        Installed = Version

                    # regex_pattern_object = re.compile("change date") # will not work missing first word
                    # match_object = re.search(regex_pattern_object,DetectedSoftware)
                    # if match_object:
                    #     DetectedSoftware = DetectedSoftware.replace("change date", "") # will not work missing first word
                    #     # get date/time
                    #     Installed = dateutil.parser.parse(DetectedSoftware)
                    #     Installed = Installed.date()
                    #     DetectedSoftware = "change date" # will not work missing first word
                    #     Version = "1.0"

                    # element to match
                    temp_match_element = [DetectedSoftware, Version]

                    # parse specific customer
                    abbr_dict_specific = indexed_config[Customer]
                    parse_configuration(abbr_dict_specific, temp_match_element, ThingName, Customer, HostName, data_dict, Installed)

                    # parse generic customer
                    abbr_dict_generic = indexed_config["generic"]
                    parse_configuration(abbr_dict_generic, temp_match_element, ThingName, Customer, HostName, data_dict, Installed)

            if log_line_count % 10000 == 0:
                print("records processed", log_line_count)
            log_line_count += 1

print("total records", log_line_count)

if len(customers_list_active) == 0:
    raise ValueError("number of recognized customers is zero")

customers_list_active = list(customers_list_active)
customers_list_active = sorted(customers_list_active)

print("generating excel reports")
with pandas.ExcelWriter(excel_multiple_sheets) as writer:
    for i in range(len(customers_list_active)):
        print("exporting to excel customer: ", customers_list_active[i])

        # get dictionary/data_frame for single customer
        single_customer_data_dict = {key: value for (key,value) in data_dict.items() if key[1] == customers_list_active[i]}
        data_frame = create_data_frame(single_customer_data_dict)

        # export to list of files (without writer)
        single_customer_file_name = reports_dir_single_customers + os.sep + customers_list_active[i] + ".xlsx"
        data_frame.to_excel(single_customer_file_name)

        # export to file with sheets (with writer)
        sheet_name = customers_list_active[i]
        sheet_name = sheet_name[:31]
        data_frame.to_excel(writer, sheet_name=sheet_name)

# export to master table
print("exporting master table to excel")
data_frame = create_data_frame(data_dict)
data_frame.to_excel(excel_master_table)

print("formating excel reports")
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
for subdir, dirs, files in os.walk(reports_dir):
    for filename in files:
        print("formating file: ", filename)
        filepath = subdir + os.sep + filename
        workbook = openpyxl.load_workbook(filepath)

        # style for dates
        date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY')
        workbook.add_named_style(date_style)

        for i in range(len(workbook.worksheets)):
            sheet = workbook.worksheets[i]
            print("formating customer sheet: ", sheet.title)

            # insert table
            table_name = sheet.title.replace(" ", "")
            table_length = sheet.max_row
            table_width = get_column_letter(sheet.max_column)
            tab = openpyxl.worksheet.table.Table(displayName=table_name, ref=f'A1:{table_width}{table_length}')
            style = openpyxl.worksheet.table.TableStyleInfo(name="None", showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            sheet.add_table(tab)

            # autofit labels width
            correction = 8
            column_text_width = 12
            column_final_width = 1
            for cell in sheet[1]:
                column_label_width = len(str(cell.internal_value)) + correction
                column_letter = cell.column_letter
                if column_label_width < column_text_width:
                    column_final_width = column_text_width
                else:
                    column_final_width = column_label_width
                sheet.column_dimensions[column_letter].width = column_final_width

            sheet.column_dimensions["C"].width = 8 + correction
            sheet.column_dimensions["A"].width = 9 + correction

            max_customer_name_length = 1
            for cell in sheet["B"]:
                if len(cell.internal_value) > max_customer_name_length:
                    max_customer_name_length = len(cell.internal_value)
            sheet.column_dimensions["B"].width = max_customer_name_length + 2

            # mark duplicates
            temp_cell = sheet.cell(1,1)
            for cell in sheet['C']:
                if cell.internal_value == temp_cell.internal_value:
                    cell.fill = redFill
                    temp_cell.fill = redFill
                temp_cell = cell

            # center justify & style dates
            for column in sheet.iter_cols():
                for cell in column:
                    cell.alignment = Alignment(horizontal='center')
                    if isinstance(cell, datetime.datetime):
                        cell.style = date_style

            # freeze panes
            cell = sheet['D2']
            sheet.freeze_panes = cell

        workbook.save(filepath)
print()
print("all reports finished")
sleep(3)
